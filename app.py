# -*- coding: utf-8 -*-
"""
Created on Thu Sep 17 15:54:14 2020

@author: niels
"""
from convert_files import convert_pdfs
from scroll_pdf import pdf_canvas
from produce_feedback import excel_feedback
import tkinter as tk
import re
import pandas as pd
import openpyxl as oxl
from win32com.client import Dispatch
import os
import shutil
# import traceback

def c(col,row):
    abc = list("0ABCDEFGHIJKLMNO")
    return abc[col]+str(row)

def sstr(s):
    return s.encode('ascii', errors='ignore').decode('utf-8')

class pre_interface(tk.Frame):
    '''The interface that helps to make the names.txt and questions.txt file.'''
    
    def __init__(self, parent, names, sheets, path):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.names = names
        self.sheets = sheets
        self.path = path
        
        sheet_lb    = tk.Label(self, justify = tk.LEFT, text="Questions (seperated by enter)\nformat :\t\t\'x[abc]) title'\nexample :\t'2a) Determine eigenvalues'")
        sheet_lb.grid(column = 0, row = 0, sticky = "nsew")
        self.sheet = tk.Text(self, height = 40, width = 80)
        self.sheet.grid(column = 0, row = 1, sticky = "nsew")
        self.sheet.insert(tk.END, self.sheets)
        
        name_lb     = tk.Label(self, text="Names (seperated by enter)")
        name_lb.grid(column = 1, row = 0, sticky = "nsew")
        self.name = tk.Text(self,height = 40, width = 30)
        self.name.grid(column = 1, row = 1, sticky = "nsew")
        self.name.insert(tk.END, self.names)
        
        self.enter = tk.Button(self, text="Finish", command = self.enter_layout, relief = tk.GROOVE)
        self.enter.grid(column = 0, row = 2, sticky = "nsew", columnspan = 2)
        self.enter.bind("<Enter>", self.enter_enter)
        self.enter.bind("<Leave>", self.enter_leave)
    
    def enter_layout(self):
        self.names = self.name.get("1.0",'end-1c').encode('utf-8').strip().split("\n")
        self.sheets = self.sheet.get("1.0",'end-1c').strip().split("\n")
        self.names.sort()
        
        self.write_names_and_sheets(self.names, self.sheets, self.path)
        self.root.destroy()
    
    def enter_enter(self, e):
        self.enter['background'] = 'green'
    
    def enter_leave(self, e):
        self.enter['background'] = 'SystemButtonFace'
        
    def write_names_and_sheets(self, names, sheets, path):
        file = path+"questions.txt"
        try:
            with open(file, 'w') as f:
                for sheet in sheets:
                    f.write(sheet,'\n')
        except:
            print(file+" error")
            
        file = path+"names.txt"
        try:
            with open(file, 'w') as f:
                for name in names:
                    f.write(name,'\n')
        except:
            print(file+" error")
        pass

class gui(tk.Frame):
    '''This contains the actual interface while grading
    the buttons, pdf's, connection to excel scoring sheet 
    and keyboard instructions for fast grading'''
    
    def __init__(self,
                 parent,
                 names,
                 sheets,
                 sheet_names,
                 path,
                 file_scoring,
                 answer_pdf):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.names = names
        self.sheets = sheets
        self.sheet_names = sheet_names
        self.path = path
        self.file_scoring = file_scoring
        self.answer_pdf = answer_pdf
        
        self.RW, self.RH      = 640,240;
        self.Rw = lambda f : int(f*self.RW); 
        self.Rh = lambda f : int(f*self.RH);
        self.PADX, self.PADY  = 4,4;
        
        self.sheet_txt   = "question:\n"
        self.sheet_lb    = tk.Label(self,text=self.sheet_txt)
        self.name_txt    = "person:\n"
        self.name_lb     = tk.Label(self,text=self.name_txt)
        
        self.name_forward      = tk.Button(self,text="next person", command = self.next_person)
        self.name_backward     = tk.Button(self,text="prev person", command = self.prev_person)
        self.sheet_forward     = tk.Button(self,text="next question", command = self.next_sheet)
        self.sheet_backward    = tk.Button(self,text="prev question", command = self.prev_sheet)
        
        self.new_option  = tk.Button(self,text="NEW GRADING OPTION\n(max 7)", command = self.new_grade_option)
        self.new_option_make = False
        self.no_description = tk.Text(self,height = 6, width = 4)
        self.no_deduction   = tk.Entry(self)

        self.new_remark_make = False
        
        #data = x,y,colspan,rowspan
        #or
        #data = x,y,(chars width, lines height)
        self.WL = self.Rw(.16) #wraplength
        self.grid_items = {
            self.sheet_lb    : (0,0,(self.WL,1)),
            self.sheet_backward    : (1,0,(self.WL,1)),
            self.sheet_forward     : (2,0,(self.WL,1)),
            self.name_lb     : (3,0,(self.WL,1)),
            self.name_backward     : (4,0,(self.WL,1)),
            self.name_forward      : (5,0,(self.WL,1)),
            self.new_option  : (1,1,(self.WL,1)),
            }
        
        self.no_items = {
            self.no_description : (6,0),
            self.no_deduction   : (7,0),
            }
        
        
        self.initial = True
        self.sheet = self.sheets[0]
        self.name = self.names[0]
        self.options = []
        self.get_sheet()
        self.options[0].focus()
        self.bind_keys()
        self.initial = False
        
        self.new_grade_option()
        self.new_grade_option()
        
        self.pack()
        
        self.doc_frame = tk.Frame(self.parent)
        
        #pdf of answers
        self.answer_obj = pdf_canvas(self.doc_frame, self.answer_pdf, zoom = 5)
        self.answer_frame,self._answer = self.answer_obj.get_canvas()
        self.answer_frame.pack(side = tk.RIGHT, anchor = tk.NE)
        
        #pdf of person
        self.zoom = 5
        pdf0 = self.path + "\\" + sstr(self.name) + ".pdf"
        self.pdf_obj = pdf_canvas(self.doc_frame, pdf0, zoom = self.zoom)
        self.pdf_obj.func_image = self.pdf_obj.get_jpg #for faster pdf loading
        self.pdf_frame,self._pdf = self.pdf_obj.get_canvas()

        self.pdf_frame.pack(side = tk.LEFT, anchor = tk.NW)
        self.pdf_obj.scroll_bar.pack_forget()
        self.pdf_obj.scroll_bar.pack(side = tk.RIGHT, fill = tk.Y)
        self.pdf_yscroll = [0 for _ in names]
        
        self.doc_frame.pack(side = tk.BOTTOM)
    
    def refresh_grid(self, dic):
        for item,data in dic.items():
            #data = x,y,colspan,rowspan
            #or
            #data = x,y,(chars width, lines height)
            item.grid(column = data[0], row = data[1], sticky = "nsew")
            item.grid_configure(padx=self.PADX, pady=self.PADY)
            self.columnconfigure(data[0], minsize = self.WL+self.Rw(.1))
            self.rowconfigure(data[1], minsize = self.Rh(.5))
            if len(data) > 3:
                item.grid_configure(columnspan = data[2], rowspan = data[3])
            #check the text size
            elif len(data) > 2:
                item.configure(wraplength = str(data[2][0])+'p')
                if len(item["text"]) > data[2][0]*data[2][1]:
                    new_text = item["text"][:-3]+"..."
                    item.configure(text = new_text)
                
    def option_toggle(self, i):
        #i is button index in options
        global name,sheet
        try:
            wb = oxl.load_workbook(self.file_scoring)
            ws = wb[self.sheet]
                
            current_value = ws[c(3+i,3+self.names.index(self.name))].value
            if current_value == None:
                ws[c(3+i,3+self.names.index(self.name))] = 'x'
                self.options[i].config(relief = "sunken",bg="yellow")
            elif current_value == 'x':
                ws[c(3+i,3+self.names.index(self.name))] = None
                self.options[i].config(relief = "raised",bg="white")
            else:
                ws[c(3+i,3+self.names.index(self.name))] = None
                self.options[i].config(relief = "raised",bg="white")
                print("weird value -->  reset to None")
        finally:
            wb.save(self.file_scoring)
            wb.close()
    
    def get_sheet(self):
        try:
            wb = oxl.load_workbook(self.file_scoring)
            ws = wb[self.sheet]
            
            if not self.initial:
                for option in self.options:
                    del self.grid_items[option]
                    option.destroy()
            
            #set person and sheet button text
            self.name_lb["text"] = self.name_txt + sstr(self.name)
            self.sheet_lb["text"] = self.sheet_txt + self.sheet_names[self.sheet]
            
            #get options
            self.options = [];
            for col in range(3,10):
                description = ws[c(col,1)].value
                if description != None:
                    option = tk.Button(self, text=description, bg="white",
                                       command = lambda x=col-3: self.option_toggle(x))
                    self.options.append(option)
                    self.grid_items[option] = (col-3,1,(self.WL,1))
                    self.grid_items[self.new_option] = (col-2,1,(self.WL,1))
                    self.refresh_grid(self.grid_items)
    
            #and get their current values
            self.get_current_options()
        finally:
            wb.close()
        
        self.update_screensize()
    
    def get_current_options(self):
        try:
            wb = oxl.load_workbook(self.file_scoring)
            ws = wb[self.sheet]
            
            #get options
            for col2 in range(3,len(self.options)+3):
                val = ws[c(col2,self.names.index(self.name)+3)].value
                if val == 'x':
                    self.options[col2-3].config(relief = "sunken",bg="yellow")
                elif val == None:
                    self.options[col2-3].config(relief = "raised",bg="white")
                else:
                    self.options[col2-3].config(relief = "raised",bg="white")
                    ws[c(col2,self.names.index(self.name)+3)] = None
                    print("This cell has weird value --> reset to None")
        finally:
            wb.close()
    
    def new_grade_option(self):
        if self.new_remark_make:
            return
        if self.new_option_make:
            description = self.no_description.get("1.0",tk.END).strip()
            deduction = self.no_deduction.get().strip()
            
            self.no_description.grid_remove()
            self.no_deduction.grid_remove()
            
            if description != "" and deduction != "":
                try:
                    wb = oxl.load_workbook(self.file_scoring)
                    ws = wb[self.sheet]
                    
                    for col in range(3,10):
                        val = ws[c(col,1)].value
                        if val == None:
                            option = tk.Button(self, text=description, bg="white"
                                               ,command = lambda x=col-3: self.option_toggle(x))
                            self.options.append(option)
                            self.grid_items[option] = (col-3,1,(self.WL,1))
                            self.grid_items[self.new_option] = (col-2,1,(self.WL,1))
                            self.refresh_grid(self.grid_items)
                            
                            ws[c(col,1)] = description
                            ws[c(col,2)] = round(float(deduction),2)
                            
                            print("Added new option")
                            
                            break
                        elif col == 9:
                            print("Max options (7) reached")
                            break
                finally:
                    wb.save(self.file_scoring)
                    wb.close()
            
            self.no_description.delete("1.0",tk.END)
            self.no_deduction.delete(0,tk.END)
            self.new_option_make = False
        else:
            self.refresh_grid(self.no_items)
            self.no_description.focus()
            self.new_option_make = True
        
        self.update_screensize()
    
    def new_remark(self):
        if self.new_option_make:
            return
        if self.new_remark_make:
            remark = self.no_description.get("1.0",tk.END).strip()
            self.no_description.grid_remove()
            
            try:
                wb = oxl.load_workbook(self.file_scoring)
                ws = wb[self.sheet]
                
                ws[c(12,self.names.index(self.name)+3)] = remark           
            finally:
                wb.save(self.file_scoring)
                wb.close()
            
            self.no_description.delete("1.0",tk.END)
            self.new_remark_make = False
        else:
            self.refresh_grid({k:self.no_items[k] for k in self.no_items if k == self.no_description})
            
            try:
                wb = oxl.load_workbook(self.file_scoring)
                ws = wb[self.sheet]
                
                val = ws[c(12,self.names.index(self.name)+3)].value  
                if val != None:
                    self.no_description.insert(1.0, val)
            finally:
                    wb.save(self.file_scoring)
                    wb.close()     
                    
            self.no_description.focus()
            self.new_remark_make = True
        
        self.update_screensize()
    
    def set_pdf(self):
        file = self.path + "\\" + sstr(self.name) + ".pdf"
        
        self.pdf_frame,self._pdf = self.pdf_obj.change_canvas(file)
        i = self.names.index(self.name)
        self.pdf_obj.canvas.yview_moveto(self.pdf_yscroll[i])
    
    def save_pdf_scroll(self):
        i = self.names.index(self.name)
        self.pdf_yscroll[i] = self.pdf_obj.canvas.yview()[0]
    
    def change_to_person(self, i):
        self.save_pdf_scroll()
        self.name = self.names[i]
        self.name_lb["text"] = self.name_txt + self.name
        self.get_current_options()
        self.set_pdf()
    
    def next_person(self):
        if self.names.index(self.name) == len(self.names)-1:
            i = 0
        else:
            i = self.names.index(self.name)+1
        self.change_to_person(i)
    
    def prev_person(self):
        if self.names.index(self.name) == 0:
            i = len(self.names)-1
        else:
            i = self.names.index(self.name)-1
        self.change_to_person(i)
    
    def next_sheet(self):
        if self.sheets.index(self.sheet) == len(self.sheets)-1:
            print("this is the last sheet")
        else:
            self.sheet = self.sheets[self.sheets.index(self.sheet)+1]
            self.sheet_lb["text"] = self.sheet_txt + self.sheet
            self.get_sheet()
    
    def prev_sheet(self):
        if self.sheets.index(self.sheet) == 0:
            print("this is the first sheet")
        else:
            self.sheet = self.sheets[self.sheets.index(self.sheet)-1]
            self.sheet_lb["text"] = self.sheet_txt + self.sheet
            self.get_sheet()
    
    def update_screensize(self):
        ww, hh = self.grid_size()
        ww = (ww) * self.Rw(.26) + 2*self.PADX
        hh = (hh-int(not self.new_option_make)+int(self.initial)) * self.Rh(.5) + 2*self.PADY
        # self.parent.geometry(str(ww) + 'x' + str(hh))
    
    def _new_option(self,e):
        #press n for new options
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        self.new_grade_option()
        
    def _new_remark(self,e):
        #press m for edit remark
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        self.new_remark()
        
    def _next(self,e):
        #next focus
        current = self.parent.focus_get()
        if current == self.no_description or current == self.no_deduction:
            return
        elif current in self.options[:-1]:
            self.options[self.options.index(current)+1].focus()
        else:
            self.options[0].focus()
            
    def _prev(self,e):
        #prev focus
        current = self.parent.focus_get()
        if current == self.no_description or current == self.no_deduction:
            return
        elif current in self.options[1:]:
            self.options[self.options.index(current)-1].focus()
        else:
            self.options[-1].focus()
        
    def _goto_next(self,e):
        if self.parent.focus_get() == self.no_description:
            if self.new_remark_make:
                self.new_remark()
                self.parent.focus()
            else:
                self.no_deduction.focus()
        elif self.parent.focus_get() == self.no_deduction:
            self.new_grade_option()
            self.parent.focus()
        elif self.name != self.names[-1]:
            self.next_person()
            self.options[0].focus()
        elif self.sheet != self.sheets[-1]:
            self.next_sheet()
            self.change_to_person(0)
            self.options[0].focus()
        else:
            print("All finished")
            self.parent.destroy()
    
    def _select_option(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        self.parent.focus_get()["command"]
            
    def _zoom_in(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        self.zoom += 1
        self.pdf_frame,self._pdf = self.pdf_obj.change_size(self.zoom)
    
    def _zoom_out(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        self.zoom -= 1
        self.pdf_frame,self._pdf = self.pdf_obj.change_size(self.zoom)
        
    def _scroll_up(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        e.delta = 120
        self.pdf_obj.func_scroll(e)
        
    def _scroll_down(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        e.delta = -120
        self.pdf_obj.func_scroll(e)
    
    def _scroll_up_answer(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        e.delta = 120
        self.answer_obj.func_scroll(e)
        
    def _scroll_down_answer(self,e):
        if self.parent.focus_get() == self.no_description or self.parent.focus_get() == self.no_deduction:
            return
        e.delta = -120
        self.answer_obj.func_scroll(e)
    
    def bind_keys(self):
        self.parent.bind("<Return>",self._goto_next)
        self.parent.bind("j",self._prev)
        self.parent.bind("k",self._next)
        self.parent.bind("<space>",self._select_option)
        self.parent.bind("n",self._new_option)
        self.parent.bind("m",self._new_remark)
        self.parent.bind("r",self._scroll_up)
        self.parent.bind("f",self._scroll_down)
        self.parent.bind("t",self._scroll_up_answer)
        self.parent.bind("g",self._scroll_down_answer)
        # self.parent.bind("<Up>",self._zoom_in)
        # self.parent.bind("<Down>",self._zoom_out)
    pass
    
class app():
    def __init__(self, 
                 path,
                 superfolder = "",
                 folder = "",
                 file_template = "scoring_template.xlsx", 
                 file_scoring = "scores_week3.xlsx",
                 ):
        self.title = "The checker for numerical methods - by Niels Burghoorn"
        
        self.file_template = path+file_template
        self.file_scoring = path+superfolder+file_scoring
        self.path = path+superfolder
        self.folder = folder
        self.answer_pdf = input(f"Make sure you have your answers pdf-file relative in ...{self.path[-25:]}.\nIf it is 'answers.pdf' proceed, else what is its name?:\n")
        self.answer_pdf = self.path+"answers.pdf" if self.answer_pdf == "" else self.path+self.answer_pdf
        self.names = []
        self.sheets = []
        self.sheet_names = []
        
        self.gui = None
    
    def path_end_backslash(self):
        '''adds a backslash to the path if necessary'''
        if self.path[-1] != "\\":
            self.path += "\\"
        return(self.path)
    
    def get_names_and_sheets(self, path):
        file = path+"questions.txt"
        try:
            with open(file, 'r') as f:
                values = f.read().strip().splitlines()
            self.sheets = values
            print("Found questions.txt")
        except FileNotFoundError:
            print(file+" not found")
        except:
            print(file+" not good format")
            
        file = path+"names.txt"
        try:
            with open(file, 'r') as f:
                values = f.read().strip().splitlines()
            self.names = values
            print("Found names.txt")
        except FileNotFoundError:
            print(file+" not found")
        except:
            print(file+" not good format")
        pass
    
    def split_sheet_names(self):
        '''This splits the sheetnames (questions) into prefices ( e.g.:"1a)" )
        and the questions (e.g.: "find eigenvalues" )'''
        prefices = []
        for sheet in self.sheets:
            pattern = r"^[^\)]+\)"
            obj = re.match(pattern,sheet).group()
            if obj in prefices:
                i = 1
                while obj + " (" + str(i) + ")" in prefices:
                    i += 1
                prefices.append(obj + " (" + str(i) + ")")
            else:
                prefices.append(obj)
        
        self.sheet_names = dict(zip(prefices, self.sheets))
        self.sheets = prefices
    
    def create_names_and_sheets(self, path):
        qtxt = input(f"Prespecified questions.txt in ...{path[-25:]}? (yes/no):\n")
        
        if qtxt == "yes":
            self.get_names_and_sheets(path)
    
    def pre_dialog(self):
        '''starts the application
        Here the information (names and questions) are imported/specified
        Also the score sheet is copied from the template excel file'''
        
        path = self.path_end_backslash()
        self.create_names_and_sheets(path)
        
        if self.sheets == []:
            root = tk.Tk();
            root.title(self.title)
            self.gui = pre_interface(root, self.names, self.sheets, self.path)
            self.gui.pack()
            root.mainloop()
        
        self.split_sheet_names()
    
    def make_excel(self):
        '''Make excel file from template. Excel sheets correspond to sheets'''
        make_file = True
        if os.path.isfile(self.file_scoring):
            print(self.file_scoring)
            make_file = (input("File already exists. Overwrite this file? (yes/no):\n") == "yes")
            if make_file:
                os.remove(self.file_scoring)
        
        if make_file:
            writer = pd.ExcelWriter(self.file_scoring, engine = 'xlsxwriter')
            
            #make file
            try:
                pd.DataFrame([]).to_excel(writer)
                writer.save()
            finally:
                writer.close()
            
            #make sheets
            template = None
            scoring = None
            try:
                xl = Dispatch("Excel.Application")
                xl.ScreenUpdating = False
                xl.DisplayAlerts = False
                xl.EnableEvents = False
                xl.Visible = False
                
                template = xl.Workbooks.Open(self.file_template)
                scoring = xl.Workbooks.Open(self.file_scoring)
                        
                def create_question_sheets(sheets):
                    template_sheet = template.Worksheets("template")
                    template_sheet.Copy(Before=scoring.Worksheets(1))
                    
                    for sheet in self.sheets:
                        scoring_sheet = scoring.Worksheets(1)
                        scoring_sheet.Copy(Before=scoring.Worksheets(self.sheets.index(sheet)+2))
                        scoring_sheet = scoring.Worksheets("template (2)")
                        scoring_sheet.Name = sheet
                
                create_question_sheets(self.sheets)
                
                scoring.Worksheets("Sheet1").delete
            finally:
                
                template.Close(SaveChanges = False)
                scoring.Close(SaveChanges = True)
                
            
            #fill sheets
            try:
                wb = oxl.load_workbook(self.file_scoring)
                for sheet in wb.sheetnames[1:]:
                    wb[sheet][c(1,1)] = self.sheet_names[sheet]
                    for row in range(3,23):
                        if len(self.names) == row-3:
                            break
                        wb[sheet][c(1,row)] = names[row-3]
            except Exception as e:
                print("Error could not fill sheets: ", e)
            finally:
                wb.save(self.file_scoring)
                wb.close()
    
    def run(self):
        root = tk.Tk();
        root.title(self.title)
        self.gui = gui(root, self.names, self.sheets, self.sheet_names, 
                      self.path+self.folder, self.file_scoring, self.answer_pdf)
        root.mainloop()
        
def run_convert():
    global names,convert_obj
    convert_obj = convert_pdfs(path+superfolder+folder)
    names = convert_obj.convert()
    
    with open(path+superfolder+"names.txt",'w',encoding='utf-8') as f:
        for name in names:
            f.write(name+"\n")

def run_app():
    global app
    app = app(path,superfolder,folder,file_template,file_scoring)
    app.pre_dialog()
    app.make_excel()
    app.run()

def run_feedback():
    global fb
    fb = excel_feedback(path+superfolder, file_scoring)
    fb.make_feedback()
    feedback_folder = path+superfolder+"feedback_files"
    if os.path.isdir(feedback_folder):
        shutil.rmtree(feedback_folder)
    os.mkdir(feedback_folder)
    fb.write_txts(feedback_folder+"\\")
    
if __name__ == "__main__":
    TIMELINESS = True
    
    path = r"C:\Users\niels\OneDrive\OneDriveDocs\TA\Numerical Methods\student results" + "\\"
    file_template = "scoring_template_timeliness.xlsx" if TIMELINESS else "scoring_template.xlsx" 
    file_scoring = "scores_intermediate_assignment4.xlsx"
    superfolder = r"intermediate_assignment4" + "\\"    
    folder = r"Assignment 4 Download 08 October, 2020 1126"
    names = None
    
    run_convert()
    run_app()
    run_feedback()
    
    